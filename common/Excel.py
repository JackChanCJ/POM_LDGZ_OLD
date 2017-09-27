# coding=utf-8
__author__ = 'JACK_CHAN'

import sys
from time import sleep
from pages.basePage import Page
from openpyxl import load_workbook


class Excel:
    """
        filename  excel文件绝对路径
        sheetname 表名
    """
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def __str__(self):
        return self.filename

    def get_cell_value(self, columnname, which_row=1):
        self.columnname = columnname
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]
        rows = ws.rows
        cols = ws.columns
        row_value = []
        col_value = []
        for i in cols:
            row_value.append(i[0].value)
        # print "取出第一行所有的列名：%s" %row_value
        if columnname in row_value:
            col_index = row_value.index(columnname)
            for i in rows:
                col_value.append(i[col_index].value)
                # print "根据列名得到索引，根据索引值获取当前列所有行的值：%s" %col_value
        else:
            print "columnname参数不在excel中，请重新赋值。"
        return col_value[which_row]


def main():
    # filename = u'D:\\Test\\POM_LDGZ_OLD\\textdata\\登录管理.xlsx'
    filename = u"D:\\01____WorkStation\PYTHON\\POM_LDGZ_OLD\\textdata\\登录管理.xlsx"
    e = Excel(filename, '登录')
    print e
    print e.get_cell_value('username')

if __name__ == '__main__':
    main()

# !/usr/bin/env python
# _*_coding:utf-8_*_
'''
for i in range(1,10):
        for j in range(1,i+1):
                print('%s*%s=%s'%(j,i,j*i),end=' ')
        print('\t')
'''
import xlrd, xlwt


# 读取excel数据
def a():
    book = xlrd.open_workbook("D:\\wangli.xlsx")  # 打开excel
    sheet = book.sheet_by_index(0)  # 获取当前页签
    nrows = sheet.nrows  # 总行数
    ncols = sheet.ncols  # 总列数
    row_data = sheet.row_values(0)  # 读取第一行数据，列表形式展现
    col_data = sheet.col_values(1)  # 读取第二列数据，列表形式展现
    cell_data = sheet.cell_value(1, 1)  # 读取单元格数据
    cell_data1 = sheet.cell(0, 0)  # 读取单元格数据和类型
    # 读取整个excel数据
    for i in range(0, nrows):
        for j in range(0, ncols):
            cell_data2 = sheet.cell_value(i, j)
            print(cell_data2)


# 写入excel数据
book = xlwt.Workbook(encoding='utf-8')
sheet = book.add_sheet('wangli1', cell_overwrite_ok=True)  # 可重新写入
sheet.write(0, 0, 'test1')  # 写入数据
sheet.write(0, 1, '哈哈')
sheet.write(0, 2, '521')
book.save('D:\\wangli1.xls')  # 保存excel
