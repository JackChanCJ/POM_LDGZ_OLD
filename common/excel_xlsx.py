# coding=utf-8
__author__ = 'JACK_CHAN'

import os
from time import sleep
from pages.basePage import Page
from openpyxl import Workbook
from openpyxl import load_workbook
import sys

reload(sys)
sys.setdefaultencoding('utf-8')


class Excel(object):
    """
        filename  excel文件绝对路径
        sheetname 表名
    """
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def __str__(self):
        return self.filename

    def get_cell_value(self, columnname, which_row=1):
        self.columnname = columnname
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]
        rows = ws.rows
        cols = ws.columns
        row_value = []
        col_value = []
        for i in cols:
            row_value.append(i[0].value)
        # print "取出第一行所有的列名：%s" %row_value
        if columnname in row_value:
            col_index = row_value.index(columnname)
            for i in rows:
                col_value.append(i[col_index].value)
                # print "根据列名得到索引，根据索引值获取当前列所有行的值：%s" %col_value
        else:
            print "columnname参数不在excel中，请重新赋值。"
        return col_value[which_row]

    def write_by_cell(self, columnname, which_row=1):
        self.columnname = columnname
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]

        pass









def main():
    cwd = os.path.abspath('..')
    # 注意：文中只要涉及到中文的地方一律转化为unicode
    filename = unicode(cwd, 'utf-8') + u'\\textdata\\登录管理.xlsx'
    print (filename)
    wb = load_workbook(filename)
    ws = wb[u'登录']
    mr = ws.max_row
    ml = ws.max_column
    print '最大行：', mr, '行'
    print '最大列：', ml, '列'
    print ws.values
    row_list = []
    row_tuple = ()
    #  把sheet以行的形式搜集起来，并以list的方式存储
    for row_tuple in ws.iter_rows():
        row_list.append(row_tuple)
    print row_list
    print row_list[0]
    print row_list[0][0].value
    print row_list[0][1].value


    print '----------------------'



    print ws.iter_cols()
    def x(ws_obj):
        for col_tuple in ws_obj:
            for c in col_tuple:
                return c.value













if __name__ == '__main__':
    main()


