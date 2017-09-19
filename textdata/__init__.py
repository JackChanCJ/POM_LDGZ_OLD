# coding=utf-8
__author__ = 'JACK_CHAN'

import sys, openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


reload(sys)
sys.setdefaultencoding("utf-8")

def read_excel_by_cellname(filename,
                            sheet_name,
                            cell_num):
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name(sheet_name)
    value = ws[cell_num].value
    return value

def write_excel_by_cellname(w_value,
                            filename,
                            sheet_index,
                            cell_num):
    wb = load_workbook(filename)
    wb._active_sheet_index = sheet_index
    ws = wb.active
    ws[cell_num] = w_value
    wb.save(filename)
    return wb


class Excel():
    # 默认返回当前列名的第一个值
    def get_cell_value(self, sheetname, columnname, which_row=1,
                       filename=u"D:\\Test\\POM_LDGZ_OLD\\textdata\\登录管理.xlsx"):
        self.sheetname = sheetname
        self.columnname = columnname
        self.filename = filename
        wb = load_workbook(filename)
        ws = wb[sheetname]
        rows = ws.rows
        cols = ws.columns
        row_value = []
        col_value = []
        for i in cols:
            row_value.append(i[0].value)
        # print "取出第一行所有的列名：%s" %row_value
        if columnname in row_value:
            col_index = row_value.index(columnname)
            for i in rows:
                col_value.append(i[col_index].value)
                # print "根据列名得到索引，根据索引值获取当前列所有行的值：%s" %col_value
        else:
            print "columnname参数不在excel中，请重新赋值。"
        return col_value[which_row]


def main():
    e = Excel()
    print e.get_cell_value('登录', 'password')


if __name__ == '__main__':
    main()
