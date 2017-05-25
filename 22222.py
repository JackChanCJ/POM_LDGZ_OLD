# coding=utf-8
__author__ = 'JACK_CHAN'

import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from pages.basePage import Page

reload(sys)
sys.setdefaultencoding("utf-8")


filename = u'D:\\Test\\POM_LDGZ_OLD\\textdata\\登录管理.xlsx'

wb = load_workbook(filename)
ws = wb.get_sheet_by_name(u'登录')

print ws['A1'].value
print ws['A2'].value
